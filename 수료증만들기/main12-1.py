import pandas as pd
from openpyxl   import load_workbook

load_wb =  load_workbook(r"C:\Users\user\Downloads\Pyprojectsforme\수료증만들기\수료증명단.xlsx")
#수료증명단 엑셀파일ㅇ르 load_workbook으로 호출~

load_ws = load_wb.active
#load_workbook을 활성화~~~!

name_list = []
birthday_list = []
ho_list = []
for i in range(1,load_ws.max_row + 1) :
    name_list.append(load_ws.cell(i,1).value) # 1에있는 것들을 i의 범위만큼 namelist에 +++
    birthday_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i,3).value)
    
print(name_list)
print(birthday_list)
print(ho_list)

